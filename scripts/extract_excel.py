"""
One-off extractor: SALES AND INVENTORY.xlsx -> prisma/seed_data/seed.json

Maps the perfume-business Excel (matrix layout) onto the generic schema
(Category / Attribute / AttributeValue / Product / Variant / VariantAttribute)
described in CLAUDE.md section 4-5. Does NOT touch the database — Prisma
seed.ts reads the JSON this produces and writes the rows.

Decisions applied (CLAUDE.md section 6):
- Current on-hand stock (INVENTORY sheet) is imported.
- SALES sheet totals (period totals, no dates/customers) are imported as one
  lump "opening balance" sale in seed.ts, so the dashboard reflects existing
  sales instead of starting at zero. unitCost is left null - Excel never
  tracked cost per sale.
"""

import json
import re
from pathlib import Path

import openpyxl

SOURCE = Path(r"C:\Users\Reinhard Ramirez BTG\Downloads\SALES AND INVENTORY.xlsx")
OUT = Path(__file__).resolve().parent.parent / "prisma" / "seed_data" / "seed.json"

# Size column -> (label, default price in PHP). Column order matches the
# Excel header row B:H on both sheets.
SIZE_COLUMNS = [
    ("50 ML", 300),
    ("30 ML", 200),
    ("100 ML (type A)", 500),
    ("100 ML (type B)", 450),
    ("Scraps", 200),
    ("10 ML", 100),
    ("10 ML (LP)", 80),
]

MAIN_PRODUCT_ROWS = range(2, 19)  # rows 2-18 inclusive (1-indexed incl header)

# 50 ML / 10 ML oil-concentration sub-tables: (first_product_row, size_label)
# Product rows sit directly below the "20/25/.../35" concentration header row.
# NOTE: row offsets differ between sheets - SALES has one extra blank row
# before each "NN MLs only" section compared to INVENTORY.
OIL_SUBTABLES_BY_SHEET = {
    "INVENTORY": [
        (23, "50 ML"),  # header row 22, products rows 23-27
        (33, "10 ML"),  # header row 32, products rows 33-37
    ],
    "SALES": [
        (25, "50 ML"),  # header row 24, products rows 25-29
        (35, "10 ML"),  # header row 34, products rows 35-39
    ],
}
OIL_PRODUCT_COUNT = 5
# (spreadsheet column, concentration %) - column D is a spacer, no data
CONCENTRATION_COLUMNS = [(2, 20), (3, 25), (5, 30), (6, 35)]

# Oil line has its own price per (size, concentration) - confirmed from the
# "in PESO" formulas in both sheets (e.g. INVENTORY!B29 =B28*220). This is a
# separate scheme from SIZE_COLUMNS, which only applies to the main line.
# 10 ML / 20% has no formula anywhere in the workbook (column left blank by
# the client) - price is None here and must be set manually in the app.
OIL_PRICE_BY_SIZE_CONCENTRATION = {
    ("50 ML", 20): 220,
    ("50 ML", 25): 270,
    ("50 ML", 30): 300,
    ("50 ML", 35): 350,
    ("10 ML", 20): None,
    ("10 ML", 25): 100,
    ("10 ML", 30): 135,
    ("10 ML", 35): 175,
}


def clean_name(name):
    return re.sub(r"\s+", " ", str(name)).strip()


def extract_sheet(sheet):
    """Returns products list [{name, category, variants:[{size, concentration, qty}]}]
    using qty as a generic on-hand-or-sold count depending on which sheet is passed."""
    products = []
    oil_subtables = OIL_SUBTABLES_BY_SHEET[sheet.title]

    # --- Main table: rows 2-18, columns B-H = 7 sizes ---
    for row_idx in MAIN_PRODUCT_ROWS:
        row = [sheet.cell(row=row_idx, column=c).value for c in range(1, 9)]
        name = row[0]
        if name is None or not str(name).strip():
            continue
        variants = []
        for col_offset, (size_label, _) in enumerate(SIZE_COLUMNS):
            qty = row[1 + col_offset]
            if qty is None or str(qty).strip() == "":
                continue
            try:
                qty_int = int(float(qty))
            except (TypeError, ValueError):
                continue
            variants.append({"size": size_label, "concentration": None, "qty": qty_int})
        if variants:
            products.append({
                "name": clean_name(name),
                "category": "Main Decants",
                "variants": variants,
            })

    # --- Oil sub-tables: 50 ML block + 10 ML block ---
    oil_products = {}
    for first_row, size_label in oil_subtables:
        for i in range(OIL_PRODUCT_COUNT):
            row_idx = first_row + i
            name = sheet.cell(row=row_idx, column=1).value
            if name is None or not str(name).strip():
                continue
            name = clean_name(name)
            oil_products.setdefault(name, [])
            for col, concentration in CONCENTRATION_COLUMNS:
                qty = sheet.cell(row=row_idx, column=col).value
                if qty is None or str(qty).strip() == "":
                    continue
                try:
                    qty_int = int(float(qty))
                except (TypeError, ValueError):
                    continue
                oil_products[name].append({
                    "size": size_label,
                    "concentration": concentration,
                    "qty": qty_int,
                })

    for name, variants in oil_products.items():
        if variants:
            products.append({"name": name, "category": "Oil Concentration", "variants": variants})

    return products


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    categories = [
        {"name": "Main Decants"},
        {"name": "Oil Concentration"},
    ]

    size_attribute_values = [{"value": label, "sortOrder": i} for i, (label, _) in enumerate(SIZE_COLUMNS)]
    concentration_attribute_values = [
        {"value": str(c), "sortOrder": i} for i, (_, c) in enumerate(CONCENTRATION_COLUMNS)
    ]

    attributes = [
        {"name": "Size", "unit": "ML", "values": size_attribute_values},
        {"name": "Concentration", "unit": "%", "values": concentration_attribute_values},
    ]

    default_price_by_size = {label: price for label, price in SIZE_COLUMNS}

    def price_for(size, concentration):
        if concentration is not None:
            return OIL_PRICE_BY_SIZE_CONCENTRATION.get((size, concentration))
        return default_price_by_size.get(size)

    # INVENTORY -> current stock (attaches price + becomes the seeded variants)
    inventory_products = extract_sheet(wb["INVENTORY"])
    products_by_name = {}
    for p in inventory_products:
        variants = {
            (v["size"], v["concentration"]): {
                "size": v["size"],
                "concentration": v["concentration"],
                "stockQty": v["qty"],
                "price": price_for(v["size"], v["concentration"]),
            }
            for v in p["variants"]
        }
        products_by_name[p["name"]] = {"category": p["category"], "variants": variants}

    # SALES -> historical totals, used to build one opening-balance sale (see seed.ts).
    # A sold size/product may not appear in current INVENTORY (fully sold out
    # since); create a zero-stock variant for it so the sale still has
    # something to attach to instead of being silently dropped.
    sales_products = extract_sheet(wb["SALES"])
    opening_sale_lines = []  # {productName, category, size, concentration, qty, unitPrice}
    for p in sales_products:
        for v in p["variants"]:
            if v["qty"] <= 0:
                continue

            existing = products_by_name.get(p["name"])
            if existing is None:
                products_by_name[p["name"]] = {"category": p["category"], "variants": {}}
                existing = products_by_name[p["name"]]

            variant_key = (v["size"], v["concentration"])
            if variant_key not in existing["variants"]:
                existing["variants"][variant_key] = {
                    "size": v["size"],
                    "concentration": v["concentration"],
                    "stockQty": 0,
                    "price": price_for(v["size"], v["concentration"]),
                }

            opening_sale_lines.append({
                "productName": p["name"],
                "category": p["category"],
                "size": v["size"],
                "concentration": v["concentration"],
                "qty": v["qty"],
                "unitPrice": price_for(v["size"], v["concentration"]),
            })

    products = [
        {"name": name, "category": data["category"], "variants": list(data["variants"].values())}
        for name, data in products_by_name.items()
    ]

    seed = {
        "settings": [
            {"key": "currency", "value": "PHP"},
            {"key": "reinvestment_goal_min", "value": "15000"},
            {"key": "reinvestment_goal_max", "value": "20000"},
        ],
        "categories": categories,
        "attributes": attributes,
        "products": products,
        "openingSale": opening_sale_lines,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(seed, indent=2), encoding="utf-8")

    main_count = sum(1 for p in products if p["category"] == "Main Decants")
    oil_count = sum(1 for p in products if p["category"] == "Oil Concentration")
    variant_count = sum(len(p["variants"]) for p in products)
    opening_total = sum(
        line["qty"] * line["unitPrice"] for line in opening_sale_lines if line["unitPrice"]
    )
    print(f"Wrote {OUT}")
    print(f"Main products: {main_count}, Oil products: {oil_count}, Variants: {variant_count}")
    print(f"Opening sale lines: {len(opening_sale_lines)}, total: PHP {opening_total}")


if __name__ == "__main__":
    main()
